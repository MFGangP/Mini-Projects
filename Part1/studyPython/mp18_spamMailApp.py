# 대량 이메일 보내기

import smtplib # SMTP 단순 전자우편 전송 프로토콜
from email.mime.text import MIMEText # 대용량 이메일 전송 확장자 규약 (이미지, 글)
from openpyxl import load_workbook
from email.mime.multipart import MIMEMultipart # 파일 텍스트 다 가능

wbook = load_workbook('./Part1/studyPython/SpamMailList.xlsx', data_only=True)
wsheet = wbook.active # sheet1 선택

for i in range(1, wsheet.max_row+1):
    recv_mail = wsheet.cell(i,1).value
    print(recv_mail)
    try:
        # 실제 메일 전송 로직
        send_mail = '이메일 주소'
        send_pass = '비밀번호'
        smtp_name = 'smtp.naver.com'
        smtp_port = 587 # 이메일이 지나가는 고속도로
        msg = MIMEMultipart()

        msg['Subject'] = '엑셀에서 보내는 메일'
        msg['From'] = send_mail # 보내는 메일
        msg['To'] = recv_mail # 받는 메일
        msg.attach(MIMEText('보내는 내용입니다.'))

        mail = smtplib.SMTP(smtp_name, smtp_port) # 객체 생성
        mail.starttls() # 보안

        mail.login(send_mail, send_pass) # 로그인
        mail.sendmail(send_mail, recv_mail, msg.as_string())
        mail.quit()
        print(f'전송성공 : {recv_mail}')

    except Exception as e:
        print(f'수신메일 - {recv_mail}')
        print(f'전송 실패 : {e}')